from floodsystem.stationdata import build_station_list
from floodsystem.datafetcher import fetch_measure_levels
def mapping(inlst: list, lout: int, hout: int) -> list | int:
    """ Takes inlst [the INput LiST] and compresses it to have length as close to lout [Length of OUTput] as possible, 
    and times all elements by hout [Hight of OUTput.]"""
    outlst = [] # the list to be outputted
    length = len(inlst)
    if lout == 0:
        lout = length
    step = int(length / lout)
    for i in range(lout):
        outlst.append(inlst[i]*hout)
    return outlst

from openpyxl import Workbook, load_workbook

import numpy as np
import pandas as pd
import datetime
from matplotlib import pyplot as plt

# customisable inputs
stationName = 'Cam'
scaleFactor = 100000
futureDays = 1

future15Mins = 96 * futureDays  # 96 15-minutes in a day

# get raw data
stations = build_station_list()
for i in stations: 
    if i.name == stationName:
        measure_id = i.measure_id
        station = i
dt = datetime.timedelta(1)
histLvAndTime = fetch_measure_levels(measure_id, dt)
levelsLst = histLvAndTime[1] # absolute levels list

#data conditioning
rlevLst = []# relative level list
trange = station.typical_range
for i in levelsLst:
    if trange != None and i != None:
        rlevel = (i - trange[0]) / (trange[1] - trange[0])  # relative level
        rlevLst.append(rlevel)

data = mapping(rlevLst, 0, scaleFactor)
datLeng = len(data)
timeAxis = []
for i in histLvAndTime[0]:
    timeAxis.append(i.strftime("%Y/%m/%d %H:%M"))
last_data_time = histLvAndTime[0][-1]
time = last_data_time
first_data_time = histLvAndTime[0][0].strftime("%Y/%m/%d %H:%M")
for i in range(future15Mins):
    time += datetime.timedelta(0,0,0,0,15)
    timeAxis.append(time.strftime("%Y/%m/%d %H:%M"))


# init excel
filename = "data_try.xlsx"

# Writing dates with pandas
timeAxis = pd.DataFrame({'Sample Times':pd.date_range(first_data_time, periods=datLeng+future15Mins, freq=datetime.timedelta(0,0,0,0,15)) })
writer = pd.ExcelWriter(filename, engine='xlsxwriter')
timeAxis.to_excel(writer, sheet_name='test', index=False)
workbook  = writer.book
worksheet = writer.sheets['test']
formatdict = {'num_format':'yyyy/mm/dd hh/mm'}
fmt = workbook.add_format(formatdict)
worksheet.set_column('A:A', None, fmt)
writer.save()
writer.close()

# Writing water level with openpyxl
workbook = load_workbook(filename=filename)
sheet = workbook.active
for i in range(1, datLeng):
    sheet.cell(row=i+1, column=2).value = data[-i]
sheet.cell(row=datLeng+1, column=2).value = data[0]

# forecasting with openpyxl
for i in range(2, future15Mins+2):
    # predicted value
    sheet.cell(row = datLeng+i, column = 3).value = f'=FORECAST.ETS(A{datLeng+1},$B$1:$B${datLeng},$A$1:$A${datLeng},0,1)'
    # upper bound
    sheet.cell(row = datLeng+i, column = 4).value = f'=FORECAST.ETS(A{datLeng+1},$B$1:$B${datLeng},$A$1:$A${datLeng},0,1)+FORECAST.ETS.CONFINT(A{datLeng+1},$B$1:$B${datLeng},$A$1:$A${datLeng},0.97,0,1)'
    # lower bound
    sheet.cell(row = datLeng+i, column = 5).value = f'=FORECAST.ETS(A{datLeng+1},$B$1:$B${datLeng},$A$1:$A${datLeng},0,1)-FORECAST.ETS.CONFINT(A{datLeng+1},$B$1:$B${datLeng},$A$1:$A${datLeng},0.97,0,1)'

workbook.save(filename=filename)

