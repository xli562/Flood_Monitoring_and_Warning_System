from floodsystem.stationdata import build_station_list
from floodsystem.datafetcher import fetch_measure_levels
def mapping(inlst: list, lout: int, hout: int) -> list | int:
    """ Takes inlst [the INput LiST] and compresses it to have length as close to lout [Length of OUTput] as possible, 
    and times all elements by hout [Hight of OUTput.]"""
    outlst = [] # the list to be outputted
    length = len(inlst)
    step = int(length / lout)
    for i in range(lout):
        outlst.append(inlst[i]*hout)
    return outlst

def numToABC(num: int) -> str:
    ABClst = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG']
    return ABClst[num-1]

from openpyxl import load_workbook
import subprocess, os
import pyautogui

import numpy as np
import pandas as pd
import datetime
import time
from matplotlib import pyplot as plt

def predict(stations: list, scaleFactor: int=100000, pastDays: int | float=1, futureDays: int=1, stationsCount: int=5, usePrevExcel: bool = False, seasonal: bool = False) -> dict:
    """Takes a list of MonitoringStation objects, 
    a scalefactor (the prediction model is not accurate for small numbers or data with small variation)
    the days into the past that the prediction is based on
    the number of days into the future where water level is to be predicted
    the number of stations to be predicted
    usePrevExcel: whether the excel from the previous run should be used. Can save time.

    Outputs a dictionary with station names as keys and lists of predicted **SCALED** relative river levels as words.
    
    **MAKE SURE MICROSOFT EXCEL IS INSTALLED BEFORE USING THIS FUNCTION. 
    DO NOT USE THE KEYBOARD OR THE MOUSE WHILE THE FUNCTION IS RUNNING**"""

    past15Mins = 96 * pastDays      # 96 15-minute intervals in a day
    future15Mins = 96 * futureDays  # 96 15-minute intervals in a day
    stations = stations[:stationsCount]
    seasonInt = int(seasonal)
    # =======Getting and conditioning the data=======
    pastDatas = []                # List of past datas
    for ii in stations:                 # ii is a MonitoringStation object in the 'stations' list
        # ===Getting the raw data===
        measureID = ii.measure_id       # The measure ID of the station
        station = ii                    # A more appropriate name for ii, and in case ii gets changed.
        dt = datetime.timedelta(pastDays)                     # The numbers of days into the past, whose water level data the prediction will be based on
        historyLevelAndTime = fetch_measure_levels(measureID, dt)     # A list of the measurement time and the history water level
        levelsLst = historyLevelAndTime[1] # absolute levels list     # A list of history water levels recorded at the station
        # ===Data conditioning===
        rlevLst = []                    # A list of the relative water levels

        # Calculating relative water levels
        trange = station.typical_range 
        for i in levelsLst:
            if trange != None and i != None:
                rlevel = (i - trange[0]) / (trange[1] - trange[0])  # relative level
                rlevLst.append(rlevel)

        pastDatas.append(rlevLst)

    # Getting the smallest size of all the datas
    datLeng = len(pastDatas[0])       # Initializing the datLeng [leength of past data] list
    for i in pastDatas:
        leni = len(i)           # A temporary vatiable to store the length of i
        if leni < datLeng:
            datLeng = leni

    # Compressing all data to (hopefully) the same length   <- might cause problems if all data are not of the same length
    tempLst = []    # A temporary list
    for i in pastDatas:
        tempLst.append( mapping(i, datLeng, scaleFactor) )
    pastDatas = tempLst
    # =======Constructing the Time axis (includes both past and future measurement times)=======
    lastDataTime = historyLevelAndTime[0][-1]             # The last instant when history data is recorded
    lastTime = lastDataTime                         # The last item on the Time axis
    firstDataTime = historyLevelAndTime[0][0].strftime("%Y/%m/%d %H:%M")   # Time of the first data point
    timeAxis = pd.DataFrame({'Water Level Sampling Times':pd.date_range(firstDataTime, periods=datLeng+future15Mins, freq=datetime.timedelta(0,0,0,0,15)) })

    # =======Using Excel for the prediction=======
    filename = "PredictionSheet.xlsx"

    if usePrevExcel == False:
        # ===Writing dates with pandas===
        writer = pd.ExcelWriter(filename, engine='xlsxwriter')
        timeAxis.to_excel(writer, sheet_name='FloodWarning', index=False)
        workbook  = writer.book
        worksheet = writer.sheets['FloodWarning']
        formatdict = {'num_format':'yyyy/mm/dd hh/mm'}
        fmt = workbook.add_format(formatdict)
        worksheet.set_column('A:A', None, fmt)
        writer.save()
        writer.close()

        # ===Writing water levels' history data with openpyxl===
        workbook = load_workbook(filename=filename)
        sheet = workbook.active
        counter = 0
        for pastData in pastDatas:
            counter += 1
            for i in range(1, datLeng):
                sheet.cell(row=i+1, column=1+counter).value = pastData[-i]
            sheet.cell(row=datLeng+1, column=1+counter).value = pastData[0]

        # ===Forecasting with openpyxl===
            for i in range(2, future15Mins+2):
                # predicted value
                ABC = numToABC(counter+1)
                sheet.cell(row = datLeng+i, column = 1+counter).value = f'=FORECAST.ETS(A{datLeng+i-1},${ABC}$1:${ABC}${datLeng},$A$1:$A${datLeng},{seasonInt},1)'
                # upper bound
                # sheet.cell(row = datLeng+i, column = 4).value = f'=FORECAST.ETS(A{datLeng+1},$B$1:$B${datLeng},$A$1:$A${datLeng},0,1)+FORECAST.ETS.CONFINT(A{datLeng+1},$B$1:$B${datLeng},$A$1:$A${datLeng},0.97,0,1)'
                # lower bound
                # sheet.cell(row = datLeng+i, column = 5).value = f'=FORECAST.ETS(A{datLeng+1},$B$1:$B${datLeng},$A$1:$A${datLeng},0,1)-FORECAST.ETS.CONFINT(A{datLeng+1},$B$1:$B${datLeng},$A$1:$A${datLeng},0.97,0,1)'

        workbook.save(filename=filename)

        # Deleting the '@' symbols
        os.startfile(filename)
        time.sleep(2)
        pyautogui.hotkey('ctrl', 'f')
        pyautogui.press('tab', presses = 5, interval = 0.02)
        pyautogui.press('right')
        pyautogui.write('@', interval = 0.02)
        pyautogui.press('tab', presses = 3, interval = 0.02)
        pyautogui.press('enter')
        time.sleep(2+(stationsCount-3)*0.2)
        pyautogui.press(['enter', 'escape'], interval = 0.1)
        pyautogui.hotkey('ctrl', 's')
        time.sleep(1.5)
        pyautogui.hotkey('alt', 'F4')

    # Reading from the excel file
    workbook = load_workbook(filename=filename, read_only=True, data_only=True)
    sheet = workbook.active
    outLst = []
    outDict = {}
    counter = 0
    for ii in range(stationsCount):
        column = numToABC(ii+2)
        for i in range(datLeng+2, datLeng+future15Mins+2): 
            outLst.append(sheet[f"{column}{i}"].value)
        outDict[stations[counter].name] = outLst
        counter += 1
    return outDict

a = predict(build_station_list(), pastDays=2, futureDays = 1, stationsCount=2, usePrevExcel=True)
print(a)
